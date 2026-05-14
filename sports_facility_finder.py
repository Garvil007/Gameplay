import re
import time
import math
import io
import json
import hashlib
import sqlite3
import os
import requests
import streamlit as st
from dotenv import load_dotenv

load_dotenv()
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OVERPASS_MIRRORS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
    "https://overpass.osm.ch/api/interpreter",
    "https://overpass.openstreetmap.fr/api/interpreter",
]
DEFAULT_OVERPASS_URL = OVERPASS_MIRRORS[0]

NOMINATIM_URL = "https://nominatim.openstreetmap.org"

_CONTACT = os.environ.get("CONTACT_EMAIL", "")
if _CONTACT:
    USER_AGENT = f"SportsFacilityFinder/1.0 ({_CONTACT})"
else:
    USER_AGENT = "SportsFacilityFinder/1.0"

HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept": "application/json",
    "Accept-Language": "en-US,en;q=0.9",
}

_log_lock = Lock()

CACHE_DB_PATH = "facility_cache.db"
CACHE_TTL_SECONDS = 7 * 24 * 3600

_cache_lock = Lock()

def _init_cache():
    with sqlite3.connect(CACHE_DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS api_cache (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                created_at INTEGER NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_created ON api_cache(created_at)")

def _cache_key(prefix, *args):
    raw = prefix + "|" + "|".join(str(a) for a in args)
    return hashlib.sha256(raw.encode()).hexdigest()

def cache_get(key, max_age_seconds=None):
    ttl = max_age_seconds if max_age_seconds is not None else CACHE_TTL_SECONDS
    with _cache_lock:
        with sqlite3.connect(CACHE_DB_PATH) as conn:
            row = conn.execute(
                "SELECT value, created_at FROM api_cache WHERE key = ?", (key,)
            ).fetchone()
            if not row:
                return None
            value, created_at = row
            if time.time() - created_at > ttl:
                return None
            try:
                return json.loads(value)
            except Exception:
                return None

def cache_set(key, value):
    with _cache_lock:
        with sqlite3.connect(CACHE_DB_PATH) as conn:
            conn.execute(
                "INSERT OR REPLACE INTO api_cache (key, value, created_at) VALUES (?, ?, ?)",
                (key, json.dumps(value), int(time.time())),
            )

def _cache_set_with_ttl(key, value, ttl_seconds):
    cache_set(key, value)

def cache_clear():
    with _cache_lock:
        with sqlite3.connect(CACHE_DB_PATH) as conn:
            conn.execute("DELETE FROM api_cache")

def cache_stats():
    if not os.path.exists(CACHE_DB_PATH):
        return 0, 0
    with _cache_lock:
        with sqlite3.connect(CACHE_DB_PATH) as conn:
            count = conn.execute("SELECT COUNT(*) FROM api_cache").fetchone()[0]
    size = os.path.getsize(CACHE_DB_PATH)
    return count, size

_init_cache()

SPORTS_CONFIG = {
    "Soccer / Football": {
        "osm_sports": ["soccer", "football"],
        "keywords": ["soccer", "football", "futbol", "fútbol", "athletic field",
                     "sports field", "multi-purpose", "multipurpose"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "tennis center", "library", "marina", "model airplane"],
        "facility_label": "Soccer Field",
        "label_variants": {
            "soccer_football": "Soccer/Football Field",
            "football_only": "Football Field",
            "multi": "Multi-Purpose Field (Soccer)",
        },
        "section_keywords": ["park", "field", "sports", "recreation"],
        "min_pitch_length_ft": 180,
    },
    "Baseball / Softball": {
        "osm_sports": ["baseball", "softball"],
        "keywords": ["baseball", "softball", "diamond", "little league",
                     "ball field", "ballfield", "tee ball", "t-ball"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "tennis center", "library", "marina"],
        "facility_label": "Baseball Field",
        "label_variants": {
            "softball": "Softball Field",
            "both": "Baseball/Softball Field",
        },
        "section_keywords": ["park", "field", "diamond"],
        "min_pitch_length_ft": 200,
    },
    "Basketball": {
        "osm_sports": ["basketball"],
        "keywords": ["basketball", "gym", "recreation center", "rec center",
                     "community center", "boys & girls", "boys and girls",
                     "sports centre", "ymca"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "marina", "library", "model airplane"],
        "facility_label": "Basketball Court",
        "label_variants": {
            "gym": "Gymnasium Basketball Court",
            "half": "Half Court",
            "full": "Full Court",
        },
        "section_keywords": ["park", "court", "gym", "recreation"],
        "min_pitch_length_ft": 42,
    },
    "Tennis": {
        "osm_sports": ["tennis"],
        "keywords": ["tennis", "racquet", "racket club"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "marina", "library"],
        "facility_label": "Tennis Court",
        "label_variants": {},
        "section_keywords": ["park", "court", "tennis", "club"],
        "min_pitch_length_ft": 75,
    },
    "Volleyball": {
        "osm_sports": ["volleyball", "beachvolleyball"],
        "keywords": ["volleyball", "beach volleyball"],
        "exclude": ["swim", "pool", "aqua", "skatepark", "golf", "bowling",
                    "marina", "library"],
        "facility_label": "Volleyball Court",
        "label_variants": {
            "beach": "Beach Volleyball Court",
        },
        "section_keywords": ["park", "court", "beach", "gym"],
        "min_pitch_length_ft": 55,
    },
}

_TOO_SMALL_NAME_FRAGMENTS = [
    "tot lot", "tot-lot", "toddler", "mini park", "mini-park",
    "pocket park", "dog park", "dog run", "skate park", "skatepark",
    "splash pad", "spray park", "butterfly garden", "community garden",
    "meditation garden", "memorial garden", "rose garden",
]

def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    p = math.pi / 180
    a = (math.sin((lat2 - lat1) * p / 2) ** 2 +
         math.cos(lat1 * p) * math.cos(lat2 * p) *
         math.sin((lon2 - lon1) * p / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(a))

def clean_name(name):
    if not name:
        return ""
    return re.sub(r"\s+", " ", name).strip()

def meters_to_feet(m):
    return round(m * 3.28084)

def calc_dimensions_from_bounds(bounds):
    if not bounds:
        return None, None
    min_lat = bounds.get("minlat")
    max_lat = bounds.get("maxlat")
    min_lon = bounds.get("minlon")
    max_lon = bounds.get("maxlon")
    if not all([min_lat, max_lat, min_lon, max_lon]):
        return None, None
    ns = haversine(min_lat, min_lon, max_lat, min_lon)
    mid_lat = (min_lat + max_lat) / 2
    ew = haversine(mid_lat, min_lon, mid_lat, max_lon)
    length_ft = meters_to_feet(max(ns, ew))
    width_ft = meters_to_feet(min(ns, ew))
    if length_ft < 10 or length_ft > 2000:
        return None, None
    return length_ft, width_ft

def normalize_key(name):
    n = name.lower().strip()
    for suffix in [" park", " field", " fields", " court", " courts"]:
        if n.endswith(suffix):
            n = n[:-len(suffix)].strip()
    return re.sub(r"[^a-z0-9]", "", n)

NOMINATIM_MIRRORS = [
    "https://nominatim.openstreetmap.org",
    "https://nominatim.openstreetmap.de",
]

def _nominatim_request(path, params, timeout=20):
    last_error = None
    for url in NOMINATIM_MIRRORS:
        try:
            resp = requests.get(f"{url}{path}", params=params,
                                headers=HEADERS, timeout=timeout)
            if resp.status_code == 403:
                last_error = f"{url}: 403 Forbidden (User-Agent rejected)"
                continue
            if resp.status_code == 429:
                last_error = f"{url}: 429 rate limited"
                continue
            resp.raise_for_status()
            return resp.json(), url
        except requests.exceptions.HTTPError as e:
            last_error = f"{url}: HTTP {resp.status_code}"
            continue
        except requests.exceptions.Timeout:
            last_error = f"{url}: timeout"
            continue
        except requests.exceptions.ConnectionError:
            last_error = f"{url}: connection refused"
            continue
        except Exception as e:
            last_error = f"{url}: {type(e).__name__}"
            continue

    raise RuntimeError(f"All Nominatim mirrors failed. Last error: {last_error}")

def lookup_city_bbox(city, county, state="California", country="USA", use_cache=True):
    key = _cache_key("city_bbox_v2", city, county, state, country)
    if use_cache:
        cached = cache_get(key)
        if cached is not None:
            return cached

    VALID_CLASSES = {"boundary", "place"}
    target_lower = city.lower().strip()

    def _validate(item):
        cls = item.get("class", "").lower()
        typ = item.get("type", "").lower()
        if cls not in VALID_CLASSES:
            return False
        if "county" in typ:
            return False
        display = item.get("display_name", "").lower()
        first_part = display.split(",")[0].strip()
        if target_lower not in first_part:
            return False
        if "county" in first_part and "county" not in target_lower:
            return False
        return True

    strategies = []
    base_params = {"format": "json", "limit": 5, "addressdetails": 1,
                    "polygon_geojson": 1}

    structured = {**base_params, "city": city, "country": country}
    if state:
        structured["state"] = state
    if county:
        structured["county"] = county
    strategies.append(("structured", structured))

    if county:
        strategies.append(("free-form full",
            {**base_params, "q": f"{city}, {county}, {state}, {country}"}))
    strategies.append(("free-form no county",
        {**base_params, "q": f"{city}, {state}, {country}"}))
    strategies.append(("free-form minimal",
        {**base_params, "q": f"{city}, {state}"}))

    valid_item = None
    for label, params in strategies:
        try:
            data, _ = _nominatim_request("/search", params)
            if not data:
                continue
            for item in data:
                if _validate(item):
                    valid_item = item
                    break
            if valid_item:
                break
        except RuntimeError as e:
            err = str(e)
            if "403" in err:
                st.error(
                    "🚫 **Nominatim blocked the request (HTTP 403).**\n\n"
                    "OSM's Nominatim requires a valid User-Agent with real contact "
                    "info, and blocks deployments that don't comply.\n\n"
                    "**If you deployed on Streamlit Cloud:**\n"
                    "1. Go to your app → Manage app → Settings → Secrets\n"
                    "2. Add this line (use your real email):\n"
                    "```\n"
                    'CONTACT_EMAIL = "your.email@example.com"\n'
                    "```\n"
                    "3. Save → the app restarts automatically"
                )
            else:
                st.error(f"Nominatim lookup failed: {err}")
            return None
        except Exception:
            continue

    if not valid_item:
        return None

    bbox = valid_item.get("boundingbox")
    if not bbox or len(bbox) != 4:
        return None
    try:
        min_lat, max_lat = float(bbox[0]), float(bbox[1])
        min_lon, max_lon = float(bbox[2]), float(bbox[3])
    except (ValueError, TypeError):
        return None

    lat_span = max_lat - min_lat
    lon_span = max_lon - min_lon
    lat_buf = max(lat_span * 0.08, 0.005)
    lon_buf = max(lon_span * 0.08, 0.005)

    result = {
        "min_lat": min_lat - lat_buf,
        "max_lat": max_lat + lat_buf,
        "min_lon": min_lon - lon_buf,
        "max_lon": max_lon + lon_buf,
        "match_type": valid_item.get("type", ""),
        "match_class": valid_item.get("class", ""),
        "match_display": valid_item.get("display_name", ""),
    }

    geojson = valid_item.get("geojson", {})
    if geojson:
        polygon_points = _extract_polygon_points(geojson)
        if polygon_points:
            result["polygon"] = polygon_points

    if use_cache:
        cache_set(key, result)
    return result

def _extract_polygon_points(geojson):
    geom_type = geojson.get("type", "")
    coords = geojson.get("coordinates", [])
    points = []
    try:
        if geom_type == "Polygon":
            for lon, lat in coords[0]:
                points.append((lat, lon))
        elif geom_type == "MultiPolygon":
            for poly in coords:
                for lon, lat in poly[0]:
                    points.append((lat, lon))
        else:
            return None
    except (IndexError, TypeError, ValueError):
        return None
    return points if len(points) >= 3 else None

def point_in_polygon(lat, lon, polygon):
    if not polygon or len(polygon) < 3:
        return True
    inside = False
    n = len(polygon)
    j = n - 1
    for i in range(n):
        lat_i, lon_i = polygon[i]
        lat_j, lon_j = polygon[j]
        if ((lon_i > lon) != (lon_j > lon)) and \
           (lat < (lat_j - lat_i) * (lon - lon_i) / (lon_j - lon_i + 1e-12) + lat_i):
            inside = not inside
        j = i
    return inside

def in_bbox(lat, lon, bbox):
    return (bbox["min_lat"] <= lat <= bbox["max_lat"] and
            bbox["min_lon"] <= lon <= bbox["max_lon"])

class OverpassCircuitBreaker:
    def __init__(self):
        self._lock = Lock()
        self._tripped = False
        self._reason = ""

    def is_tripped(self):
        with self._lock:
            return self._tripped

    def trip(self, reason):
        with self._lock:
            if not self._tripped:
                self._tripped = True
                self._reason = reason

    def reason(self):
        with self._lock:
            return self._reason

    def reset(self):
        with self._lock:
            self._tripped = False
            self._reason = ""

def build_overpass_queries(bbox, sport_config):
    bbox_str = f"{bbox['min_lat']},{bbox['min_lon']},{bbox['max_lat']},{bbox['max_lon']}"
    sports_pattern = "|".join(sport_config["osm_sports"])

    return {
        f"{sport_config['facility_label']}s": f"""[out:json][timeout:90];
(node["leisure"="pitch"]["sport"~"{sports_pattern}"]({bbox_str});
 way["leisure"="pitch"]["sport"~"{sports_pattern}"]({bbox_str}););out center tags bb;""",

        "Parks": f"""[out:json][timeout:90];
(node["leisure"="park"]({bbox_str});
 way["leisure"="park"]({bbox_str}););out center tags;""",

        "Schools": f"""[out:json][timeout:90];
(node["amenity"~"school|college|university"]({bbox_str});
 way["amenity"~"school|college|university"]({bbox_str}););out center tags;""",

        "Sports centres + gyms": f"""[out:json][timeout:90];
(node["leisure"="sports_centre"]({bbox_str});
 way["leisure"="sports_centre"]({bbox_str});
 node["leisure"="fitness_centre"]({bbox_str});
 way["leisure"="fitness_centre"]({bbox_str});
 node["building"="sports_hall"]({bbox_str});
 way["building"="sports_hall"]({bbox_str}););out center tags;""",

        "Recreation grounds": f"""[out:json][timeout:90];
(node["landuse"="recreation_ground"]({bbox_str});
 way["landuse"="recreation_ground"]({bbox_str}););out center tags;""",

        "Community centres": f"""[out:json][timeout:90];
(node["amenity"="community_centre"]({bbox_str});
 way["amenity"="community_centre"]({bbox_str}););out center tags;""",
    }

def query_overpass(name, query, overpass_url, status_callback, breaker,
                    use_cache=True, timeout=45):
    if use_cache:
        cache_k = _cache_key("overpass", overpass_url, query)
        cached = cache_get(cache_k)
        if cached is not None:
            with _log_lock:
                status_callback(f"  [{name}] 💾 cached ({len(cached)} elements)")
            return name, cached

    if breaker.is_tripped():
        with _log_lock:
            status_callback(f"  [{name}] ⏭️ skipped (Overpass unavailable)")
        return name, []

    mirror_name = (overpass_url.split("//")[1].split("/")[0]
                    if "//" in overpass_url else overpass_url)
    resp = None
    try:
        with _log_lock:
            status_callback(f"  [{name}] querying {mirror_name}...")
        resp = requests.post(overpass_url, data={"data": query},
                             timeout=timeout, headers=HEADERS)
        resp.raise_for_status()
        elems = resp.json().get("elements", [])
        with _log_lock:
            status_callback(f"  [{name}] ✅ {len(elems)} elements")
        if use_cache:
            cache_set(_cache_key("overpass", overpass_url, query), elems)
        return name, elems
    except requests.exceptions.HTTPError:
        status = resp.status_code if resp is not None else 0
        reason = f"HTTP {status}"
    except requests.exceptions.Timeout:
        reason = f"timeout after {timeout}s"
    except requests.exceptions.ConnectionError:
        reason = "connection refused"
    except Exception as e:
        reason = f"{type(e).__name__}"

    breaker.trip(reason)
    with _log_lock:
        status_callback(f"  [{name}] ❌ {reason} — tripping circuit breaker, "
                        f"remaining queries will skip")
    return name, []

def probe_overpass(overpass_url, timeout=5):
    probe_query = "[out:json][timeout:3];node(0,0,0.001,0.001);out count;"
    resp = None
    try:
        resp = requests.post(overpass_url, data={"data": probe_query},
                             timeout=timeout, headers=HEADERS)
        resp.raise_for_status()
        return overpass_url, True, "ok"
    except requests.exceptions.HTTPError:
        status = resp.status_code if resp is not None else 0
        return overpass_url, False, f"HTTP {status}"
    except requests.exceptions.Timeout:
        return overpass_url, False, f"timeout"
    except requests.exceptions.ConnectionError:
        return overpass_url, False, "refused"
    except Exception as e:
        return overpass_url, False, type(e).__name__

def pick_working_overpass(preferred_url, status_callback, is_local=False):
    if is_local:
        url, ok, reason = probe_overpass(preferred_url, timeout=5)
        name = url.split("//")[1].split("/")[0] if "//" in url else url
        if ok:
            status_callback(f"  ✅ {name} responsive")
            return url
        else:
            status_callback(f"  ❌ {name} {reason}")
            return None

    urls_to_try = [preferred_url]
    for mirror in OVERPASS_MIRRORS:
        if mirror != preferred_url:
            urls_to_try.append(mirror)

    status_callback(f"  Probing {len(urls_to_try)} endpoints in parallel...")

    with ThreadPoolExecutor(max_workers=len(urls_to_try)) as executor:
        futures = {executor.submit(probe_overpass, url, 5): url for url in urls_to_try}
        for future in as_completed(futures):
            url, ok, reason = future.result()
            name = url.split("//")[1].split("/")[0] if "//" in url else url
            if ok:
                status_callback(f"  ✅ {name} responsive (using this)")
                for f in futures:
                    f.cancel()
                return url
            else:
                status_callback(f"  ❌ {name} {reason}")

    return None

def fetch_overpass(bbox, sport_config, overpass_url, status_callback,
                    is_local=False, use_cache=True):
    is_local_url = ("localhost" in overpass_url or "127.0.0.1" in overpass_url
                    or is_local)

    status_callback(f"Source 1: Overpass API ({'LOCAL' if is_local_url else 'PUBLIC'})...")

    PROBE_CACHE_TTL = 300
    probe_cache_key = _cache_key("overpass_probe_ok", overpass_url)
    cached_working = None
    if use_cache:
        cached_working = cache_get(probe_cache_key, max_age_seconds=PROBE_CACHE_TTL)

    if cached_working:
        working_url = cached_working
        status_callback(f"  💾 Using cached working endpoint: "
                        f"{working_url.split('//')[1].split('/')[0]}")
    else:
        working_url = pick_working_overpass(overpass_url, status_callback,
                                              is_local=is_local_url)
        if working_url is None:
            status_callback("  ⚠️  No Overpass endpoint responded. "
                            "Skipping Overpass — using Nominatim only.")
            return []
        if use_cache:
            _cache_set_with_ttl(probe_cache_key, working_url, PROBE_CACHE_TTL)

    if working_url != overpass_url:
        status_callback(f"  ℹ️ Using fallback mirror")

    max_workers = 6 if is_local_url else 3
    status_callback(f"  Running queries ({max_workers} parallel workers)...")
    queries = build_overpass_queries(bbox, sport_config)
    results = []
    breaker = OverpassCircuitBreaker()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(query_overpass, qname, qtext, working_url,
                             status_callback, breaker, use_cache): qname
            for qname, qtext in queries.items()
        }
        for future in as_completed(futures):
            try:
                qname, elems = future.result()
                for el in elems:
                    lat = el.get("lat") or el.get("center", {}).get("lat")
                    lon = el.get("lon") or el.get("center", {}).get("lon")
                    tags = el.get("tags", {})
                    bounds = el.get("bounds")
                    if lat and lon:
                        length_ft, width_ft = calc_dimensions_from_bounds(bounds)
                        results.append({
                            "source": "overpass",
                            "name": clean_name(tags.get("name", "")),
                            "lat": lat, "lon": lon,
                            "sport": tags.get("sport", ""),
                            "leisure": tags.get("leisure", ""),
                            "amenity": tags.get("amenity", ""),
                            "building": tags.get("building", ""),
                            "tags": tags,
                            "length_ft": length_ft,
                            "width_ft": width_ft,
                        })
            except Exception as e:
                status_callback(f"  Worker error: {e}")

    if breaker.is_tripped():
        status_callback(f"  ⚠️ Overpass had issues: {breaker.reason()}")
    status_callback(f"  Overpass total: {len(results)} raw elements")
    return results

def build_nominatim_searches(city, state, sport_config):
    facility = sport_config["facility_label"].lower()
    return [
        f"{facility} {city} {state}",
        f"{facility}s {city} {state}",
        f"{sport_config['osm_sports'][0]} {city} {state}",
        f"sports field {city} {state}",
        f"park {city} {state}",
        f"high school {city} {state}",
        f"middle school {city} {state}",
        f"elementary school {city} {state}",
        f"college {city} {state}",
        f"recreation center {city} {state}",
        f"community center {city} {state}",
        f"sports complex {city} {state}",
        f"playground {city} {state}",
    ]

def fetch_nominatim(city, state, bbox, sport_config, status_callback, use_cache=True):
    status_callback("Source 2: Nominatim Search API...")
    searches = build_nominatim_searches(city, state, sport_config)
    results = []
    nominatim_broken = False

    for query in searches:
        cached_data = None
        if use_cache:
            key = _cache_key("nominatim_search", query)
            cached_data = cache_get(key)

        if cached_data is not None:
            data = cached_data
            from_cache = True
        else:
            if nominatim_broken:
                status_callback(f"  ⏭️ [{query[:50]}] skipped (Nominatim unavailable)")
                continue
            try:
                params = {
                    "q": query, "format": "json", "addressdetails": 1, "limit": 20,
                }
                data, _ = _nominatim_request("/search", params, timeout=30)
                if use_cache:
                    cache_set(_cache_key("nominatim_search", query), data)
                from_cache = False
                time.sleep(1.2)
            except RuntimeError as e:
                status_callback(f"  ❌ [{query[:50]}] {e}")
                nominatim_broken = True
                continue
            except Exception as e:
                status_callback(f"  ❌ [{query[:50]}] {type(e).__name__}")
                continue

        count = 0
        for item in data:
            lat = float(item.get("lat", 0))
            lon = float(item.get("lon", 0))
            name = item.get("display_name", "").split(",")[0]
            if lat and lon and in_bbox(lat, lon, bbox):
                results.append({
                    "source": "nominatim",
                    "name": clean_name(name),
                    "lat": lat, "lon": lon,
                    "sport": "", "leisure": item.get("type", ""),
                    "amenity": item.get("class", ""), "building": "",
                    "tags": item.get("address", {}),
                    "length_ft": None, "width_ft": None,
                })
                count += 1
        tag = "💾" if from_cache else "  "
        status_callback(f"  {tag} [{query[:50]}] -> {count} in-bbox")

    status_callback(f"  Nominatim total: {len(results)} results")
    return results

def is_confirmed_sport(entry, sport_config):
    sport = entry.get("sport", "").lower()
    return any(s in sport for s in sport_config["osm_sports"])

def is_facility(entry):
    leisure = entry.get("leisure", "").lower()
    amenity = entry.get("amenity", "").lower()
    building = entry.get("building", "").lower()
    name = entry.get("name", "").lower()
    return (leisure in ("park", "sports_centre", "fitness_centre",
                        "sports_hall", "stadium", "school", "college",
                        "university", "recreation_ground", "playground") or
            amenity in ("school", "college", "university",
                        "community_centre", "leisure", "amenity") or
            building in ("sports_hall",) or
            entry.get("source") == "nominatim" or
            any(k in name for k in ["park", "school", "college",
                                     "recreation", "field", "playground"]))

def is_too_small(entry, sport_config):
    name = entry.get("name", "").lower()
    if any(frag in name for frag in _TOO_SMALL_NAME_FRAGMENTS):
        return True

    leisure = entry.get("leisure", "").lower()
    is_confirmed = is_confirmed_sport(entry, sport_config)
    if leisure == "playground" and not is_confirmed:
        return True

    length_ft = entry.get("length_ft")
    if length_ft:
        min_ft = sport_config.get("min_pitch_length_ft", 0)
        if length_ft < min_ft:
            return True

    children = entry.get("child_pitches", [])
    if children:
        min_ft = sport_config.get("min_pitch_length_ft", 0)
        all_small = all(
            (c.get("length_ft") or 0) > 0 and (c.get("length_ft") or 0) < min_ft
            for c in children
        )
        if all_small:
            return True

    return False

def merge_and_deduplicate(all_sources, sport_config, status_callback):
    status_callback(f"Total raw entries: {len(all_sources)}")

    coord_seen = set()
    deduped = []
    for entry in all_sources:
        if entry.get("lat") and entry.get("lon"):
            ck = (round(entry["lat"], 5), round(entry["lon"], 5))
            if ck in coord_seen:
                continue
            coord_seen.add(ck)
        deduped.append(entry)
    status_callback(f"After coord dedup: {len(deduped)}")

    confirmed = []
    facilities = []
    for entry in deduped:
        if (entry.get("source") == "overpass"
                and entry.get("leisure") == "pitch"
                and is_confirmed_sport(entry, sport_config)):
            confirmed.append(entry)
        elif entry.get("source") == "overpass" and entry.get("leisure") == "pitch":
            continue
        elif is_facility(entry):
            facilities.append(entry)

    status_callback(f"Confirmed pitches: {len(confirmed)}")
    status_callback(f"Facilities: {len(facilities)}")

    fac_seen = {}
    exclude_list = sport_config["exclude"]
    for entry in facilities:
        name = entry["name"].strip()
        if not name:
            continue
        if any(k in name.lower() for k in exclude_list):
            continue
        key = normalize_key(name)
        if not key:
            continue
        if key not in fac_seen:
            fac_seen[key] = entry
        else:
            existing = fac_seen[key]
            if not existing.get("lat") and entry.get("lat"):
                entry["name"] = entry["name"] or existing["name"]
                fac_seen[key] = entry

    facility_list = list(fac_seen.values())

    _SCHOOL_PRIORITY = {
        "high school": 0, "high sch": 0, "preparatory": 0, "prep school": 0,
        "middle school": 1, "middle sch": 1, "junior high": 1, "intermediate": 1,
        "elementary": 2, "primary school": 2,
        "college": 3, "university": 3,
    }

    def _school_priority(name):
        n = name.lower()
        for kw, rank in _SCHOOL_PRIORITY.items():
            if kw in n:
                return rank
        return 99

    def _institution_stem(name):
        n = name.lower()
        for kw in ["high school", "middle school", "junior high", "elementary school",
                   "elementary", "primary school", "preparatory", "prep school",
                   "college", "university", "intermediate school", "intermediate"]:
            n = n.replace(kw, "").strip()
        return re.sub(r"[^a-z0-9]", "", n)

    SAME_CAMPUS_RADIUS = 60
    suppressed = set()
    fl = facility_list
    for i in range(len(fl)):
        if i in suppressed:
            continue
        for j in range(i + 1, len(fl)):
            if j in suppressed:
                continue
            a, b = fl[i], fl[j]
            if not (a.get("lat") and a.get("lon") and b.get("lat") and b.get("lon")):
                continue
            dist = haversine(a["lat"], a["lon"], b["lat"], b["lon"])
            if dist > SAME_CAMPUS_RADIUS:
                continue
            stem_a = _institution_stem(a.get("name", ""))
            stem_b = _institution_stem(b.get("name", ""))
            if not stem_a or not stem_b or stem_a != stem_b:
                continue
            pri_a = _school_priority(a.get("name", ""))
            pri_b = _school_priority(b.get("name", ""))
            if pri_a == 99 and pri_b == 99:
                continue
            if pri_a <= pri_b:
                suppressed.add(j)
            else:
                suppressed.add(i)

    if suppressed:
        status_callback(f"  Same-campus dedup removed {len(suppressed)} co-located duplicate(s)")
        facility_list = [fl[i] for i in range(len(fl)) if i not in suppressed]

    PROXIMITY_RADIUS = 200
    for pitch in confirmed:
        if not pitch.get("lat") or not pitch.get("lon"):
            continue
        best_fac = None
        best_dist = PROXIMITY_RADIUS + 1
        for fac in facility_list:
            if not fac.get("lat") or not fac.get("lon"):
                continue
            dist = haversine(pitch["lat"], pitch["lon"], fac["lat"], fac["lon"])
            if dist < best_dist:
                best_dist = dist
                best_fac = fac
        if best_fac:
            if "child_pitches" not in best_fac:
                best_fac["child_pitches"] = []
            best_fac["child_pitches"].append(pitch)
        else:
            name = pitch.get("name", "")
            if name:
                key = normalize_key(name)
                if key and key not in fac_seen:
                    fac_seen[key] = pitch
                    facility_list.append(pitch)
            elif pitch.get("lat"):
                pitch["name"] = f"{sport_config['facility_label']} ({pitch['lat']:.4f}, {pitch['lon']:.4f})"
                facility_list.append(pitch)

    results = []
    for fac in facility_list:
        has_pitches = len(fac.get("child_pitches", [])) > 0
        name_lower = fac.get("name", "").lower()
        is_sport_name = any(k in name_lower for k in sport_config["keywords"])
        leisure = fac.get("leisure", "").lower()
        amenity = fac.get("amenity", "").lower()
        is_field_facility = (leisure in ("park", "sports_centre", "fitness_centre",
                                         "sports_hall", "stadium", "school",
                                         "college", "playground") or
                             amenity in ("school", "college", "university",
                                         "community_centre") or
                             "recreation" in name_lower)
        if has_pitches or is_sport_name or is_field_facility:
            results.append(fac)

    results = [r for r in results if r.get("name")]

    before_size = len(results)
    results = [r for r in results if not is_too_small(r, sport_config)]
    removed_small = before_size - len(results)
    if removed_small:
        status_callback(f"  Removed {removed_small} too-small / wrong-type entries "
                        f"(tot lots, playgrounds, undersized pitches)")

    multi = sum(1 for r in results if len(r.get("child_pitches", [])) > 1)
    status_callback(f"After merge: {len(results)} facilities ({multi} multi-court/field)")
    return results

def _reverse_geocode_one(entry, target_city, nominatim_url=NOMINATIM_URL,
                           use_cache=True):
    tags = entry.get("tags", {})
    street = tags.get("addr:street", "")
    number = tags.get("addr:housenumber", "")
    if street:
        city = tags.get("addr:city", target_city)
        entry["address"] = f"{number} {street}, {city}".strip().lstrip(", ")
        entry["verified_city"] = city.lower()
        return entry, "osm_tags"

    lat_r = round(entry["lat"], 5)
    lon_r = round(entry["lon"], 5)
    cache_hit = False
    if use_cache:
        key = _cache_key("reverse_geocode", lat_r, lon_r)
        cached = cache_get(key)
        if cached is not None:
            entry["address"] = cached.get("address", target_city)
            entry["verified_city"] = cached.get("verified_city", "")
            return entry, "cached"

    try:
        params = {"lat": entry["lat"], "lon": entry["lon"], "format": "json",
                  "addressdetails": 1, "zoom": 18}
        data, _ = _nominatim_request("/reverse", params, timeout=20)
        addr = data.get("address", {})
        road = addr.get("road", "")
        house = addr.get("house_number", "")
        city = addr.get("city", addr.get("town", addr.get("village", "")))
        postcode = addr.get("postcode", "")
        state = addr.get("state", "")
        entry["verified_city"] = city.lower() if city else ""
        display_city = city if city else target_city
        if road:
            parts = []
            parts.append(f"{house} {road}".strip() if house else road)
            state_abbr = state[:2].upper() if state else ""
            parts.append(f"{display_city}, {state_abbr} {postcode}".strip())
            entry["address"] = ", ".join(parts).strip().rstrip(",")
        else:
            entry["address"] = data.get("display_name", target_city)

        if use_cache:
            cache_set(_cache_key("reverse_geocode", lat_r, lon_r), {
                "address": entry["address"],
                "verified_city": entry["verified_city"],
            })
    except Exception:
        entry["address"] = target_city
        entry["verified_city"] = ""
    return entry, "api"

def reverse_geocode_all(entries, target_city, status_callback,
                         use_local_nominatim=False, use_cache=True):
    nominatim_url = NOMINATIM_URL
    n = len(entries)
    status_callback(f"Reverse geocoding {n} facilities...")

    osm_count = 0
    cache_count = 0
    api_needed = []
    for entry in entries:
        tags = entry.get("tags", {})
        if tags.get("addr:street"):
            _reverse_geocode_one(entry, target_city, nominatim_url, use_cache)
            osm_count += 1
            continue
        if use_cache:
            lat_r = round(entry["lat"], 5)
            lon_r = round(entry["lon"], 5)
            key = _cache_key("reverse_geocode", lat_r, lon_r)
            cached = cache_get(key)
            if cached is not None:
                entry["address"] = cached.get("address", target_city)
                entry["verified_city"] = cached.get("verified_city", "")
                cache_count += 1
                continue
        api_needed.append(entry)

    if osm_count:
        status_callback(f"  Used OSM addr tags: {osm_count}")
    if cache_count:
        status_callback(f"  Used cache: {cache_count} 💾")

    if not api_needed:
        status_callback(f"  All {n} addresses resolved without API calls")
        return

    if use_local_nominatim:
        status_callback(f"  Parallel reverse geocode {len(api_needed)} (local)...")
        with ThreadPoolExecutor(max_workers=8) as executor:
            list(executor.map(
                lambda e: _reverse_geocode_one(e, target_city, nominatim_url, use_cache),
                api_needed
            ))
    else:
        status_callback(f"  Sequential reverse geocode {len(api_needed)} "
                        f"(public Nominatim, 1 req/sec)...")
        for i, entry in enumerate(api_needed, 1):
            _reverse_geocode_one(entry, target_city, nominatim_url, use_cache)
            if i % 10 == 0:
                status_callback(f"    progress: {i}/{len(api_needed)}")
            time.sleep(1.1)

    status_callback(f"  Reverse geocoding complete")

def get_city_neighborhoods(city, state, country="USA", use_cache=True):
    key = _cache_key("city_neighborhoods", city, state, country)
    if use_cache:
        cached = cache_get(key)
        if cached is not None:
            return set(cached)

    neighborhoods = {city.lower()}

    try:
        params = {
            "q": f"{city}, {state}, {country}",
            "format": "json", "limit": 10, "addressdetails": 1,
            "extratags": 1, "namedetails": 1,
        }
        data, _ = _nominatim_request("/search", params)
        for item in data:
            addr = item.get("address", {})
            for key_name in ("city", "town", "village", "suburb",
                              "neighbourhood", "borough", "city_district"):
                v = addr.get(key_name, "").lower().strip()
                if v and v == city.lower():
                    name = item.get("display_name", "").split(",")[0].lower().strip()
                    if name and name != city.lower():
                        neighborhoods.add(name)
    except Exception:
        pass

    if use_cache:
        cache_set(key, list(neighborhoods))
    return neighborhoods

def _parse_city_from_address(address):
    if not address:
        return ""

    parts = [p.strip() for p in address.split(",")]
    if len(parts) < 2:
        return ""

    for part in parts[1:]:
        part = part.strip()
        if not part:
            continue
        if re.match(r"^\d{5}(-\d{4})?$", part):
            continue
        if re.match(r"^[A-Z]{2}(\s+\d{5}(-\d{4})?)?$", part):
            continue
        if re.match(r"^\d+$", part):
            continue
        if part.lower() in ("usa", "united states", "us"):
            continue
        city_token = re.split(r"\s+[A-Z]{2}\s+\d{5}", part)[0].strip()
        if city_token:
            return city_token.lower()

    return ""

def filter_wrong_city(entries, target_city, target_state, bbox, status_callback,
                       use_cache=True):
    status_callback(f"Filtering facilities outside {target_city}...")
    target = target_city.lower().strip()
    polygon = bbox.get("polygon") if bbox else None

    valid_aliases = get_city_neighborhoods(target_city, target_state,
                                            use_cache=use_cache)

    if polygon:
        status_callback(f"  Using city polygon ({len(polygon)} points) "
                        f"as secondary check")
    else:
        status_callback(f"  No polygon available — using address + verified_city")

    if len(valid_aliases) > 1:
        sample = ", ".join(sorted(valid_aliases))[:80]
        status_callback(f"  Recognizing {len(valid_aliases)} aliases: {sample}")

    filtered = []
    removed = []

    for entry in entries:
        lat = entry.get("lat")
        lon = entry.get("lon")
        v_city = entry.get("verified_city", "").lower().strip()
        address = entry.get("address", "")

        address_city = _parse_city_from_address(address)

        if address_city:
            if address_city in valid_aliases or target in address_city or address_city in target:
                filtered.append(entry)
                continue
            else:
                removed.append(
                    f"{entry['name']} (address city: '{address_city}' ≠ '{target}')"
                )
                continue

        if polygon and lat and lon:
            if point_in_polygon(lat, lon, polygon):
                filtered.append(entry)
                continue
            else:
                removed.append(f"{entry['name']} (outside city polygon)")
                continue

        if v_city:
            if v_city in valid_aliases or target in v_city or v_city in target:
                filtered.append(entry)
            else:
                removed.append(f"{entry['name']} (verified city: '{v_city}' ≠ '{target}')")
            continue

        filtered.append(entry)

    status_callback(f"Removed {len(removed)} facilities outside {target_city}")
    status_callback(f"Kept {len(filtered)} facilities")
    return filtered, removed

def categorize(entries, sport_config, status_callback):
    exclude = sport_config["exclude"]
    entries = [e for e in entries if not any(k in e["name"].lower() for k in exclude)]

    categories = {
        "PUBLIC PARKS & RECREATION": [],
        "GYMNASIUM / INDOOR FACILITIES": [],
        "HIGH SCHOOLS": [],
        "MIDDLE SCHOOLS": [],
        "ELEMENTARY SCHOOLS": [],
        "COLLEGE": [],
        "OTHER FACILITIES": [],
    }

    for entry in entries:
        combined = (entry["name"] + " " + entry.get("address", "")).lower()
        if any(k in combined for k in ["high school", "high sch", "preparatory", "prep school"]):
            categories["HIGH SCHOOLS"].append(entry)
        elif any(k in combined for k in ["middle school", "middle sch", "junior high", "intermediate"]):
            categories["MIDDLE SCHOOLS"].append(entry)
        elif any(k in combined for k in ["elementary", "primary school"]):
            categories["ELEMENTARY SCHOOLS"].append(entry)
        elif any(k in combined for k in ["college", "university"]):
            categories["COLLEGE"].append(entry)
        elif any(k in combined for k in ["gym", "recreation center", "rec center",
                                          "community center", "boys & girls",
                                          "boys and girls", "sports centre",
                                          "sports center", "fitness", "indoor",
                                          "ymca"]):
            categories["GYMNASIUM / INDOOR FACILITIES"].append(entry)
        elif any(k in combined for k in ["park", "field", "memorial", "playground", "recreation"]):
            categories["PUBLIC PARKS & RECREATION"].append(entry)
        else:
            categories["OTHER FACILITIES"].append(entry)

    return categories

def expand_to_rows(entries, sport_config, category_name):
    rows = []
    label = sport_config["facility_label"]
    variants = sport_config["label_variants"]

    for entry in entries:
        children = entry.get("child_pitches", [])
        num = max(len(children), 1)
        name_lower = entry["name"].lower()
        is_gym = (category_name == "GYMNASIUM / INDOOR FACILITIES" or
                  "gym" in name_lower)

        if num <= 1 and not children:
            sport = entry.get("sport", "").lower()
            desc = label
            if "softball" in sport and "baseball" in sport and "both" in variants:
                desc = variants["both"]
            elif "softball" in sport and "softball" in variants:
                desc = variants["softball"]
            elif "football" in sport and "soccer" in sport and "soccer_football" in variants:
                desc = variants["soccer_football"]
            elif "football" in sport and "soccer" not in sport and "football_only" in variants:
                desc = variants["football_only"]
            elif is_gym and "gym" in variants:
                desc = variants["gym"]
            elif "multi" in name_lower and "multi" in variants:
                desc = variants["multi"]

            tags = entry.get("tags", {})
            if tags.get("lit", "") == "yes":
                desc += " (Lighted)"

            lat = entry["lat"]
            lon = entry["lon"]
            rows.append({
                "name": entry["name"],
                "description": desc,
                "address": entry.get("address", ""),
                "lat": lat,
                "lon": lon,
                "length_ft": entry.get("length_ft"),
                "width_ft": entry.get("width_ft"),
                "google_earth_url": f"https://earth.google.com/web/@{lat},{lon},50a,300d,35y,0h,0t,0r",
            })
        else:
            for i, child in enumerate(children, 1):
                tags = child.get("tags", {})
                child_sport = child.get("sport", "").lower()

                base = label
                if "softball" in child_sport and "baseball" in child_sport and "both" in variants:
                    base = variants["both"]
                elif "softball" in child_sport and "softball" in variants:
                    base = variants["softball"]
                elif "football" in child_sport and "soccer" in child_sport and "soccer_football" in variants:
                    base = variants["soccer_football"]
                elif is_gym and "gym" in variants:
                    base = variants["gym"]

                suffix_parts = []
                hoops = tags.get("hoops", "")
                if hoops == "1" and "half" in variants:
                    suffix_parts.append(variants["half"])
                elif hoops == "2" and "full" in variants:
                    suffix_parts.append(variants["full"])

                if tags.get("lit", "") == "yes":
                    suffix_parts.append("Lighted")

                desc = f"{base} {i}" if num > 1 else base
                if suffix_parts:
                    desc += f" ({', '.join(suffix_parts)})"

                child_lat = child.get("lat", entry["lat"])
                child_lon = child.get("lon", entry["lon"])
                rows.append({
                    "name": entry["name"],
                    "description": desc,
                    "address": entry.get("address", ""),
                    "lat": child_lat,
                    "lon": child_lon,
                    "length_ft": child.get("length_ft"),
                    "width_ft": child.get("width_ft"),
                    "google_earth_url": f"https://earth.google.com/web/@{child_lat},{child_lon},50a,300d,35y,0h,0t,0r",
                })
    return rows

def build_excel(categories, sport_config, city, state):
    wb = Workbook()
    ws = wb.active
    title = f"{sport_config['facility_label']}s - {city}"[:31]
    ws.title = title

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    section_font = Font(name="Arial", size=10, bold=True, color="1F3864")
    section_fill = PatternFill("solid", fgColor="B4C6E7")
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    row_fill_blue = PatternFill("solid", fgColor="D6E4F0")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 30
    ws.freeze_panes = "A2"

    NUM_COLS = 8

    headers = ["Name of Facility", "Description of Facility", "Address",
                "Latitude", "Longitude", "Length (ft)", "Width (ft)",
                "Google Earth"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    current_row = 2
    section_order = [
        "PUBLIC PARKS & RECREATION",
        "GYMNASIUM / INDOOR FACILITIES",
        "HIGH SCHOOLS",
        "MIDDLE SCHOOLS",
        "ELEMENTARY SCHOOLS",
        "COLLEGE",
        "OTHER FACILITIES",
    ]
    total = 0
    for section in section_order:
        entries = categories.get(section, [])
        if not entries:
            continue
        entries.sort(key=lambda x: x["name"].lower())
        data_rows = expand_to_rows(entries, sport_config, section)

        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=NUM_COLS)
        cell = ws.cell(row=current_row, column=1, value=section)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border
        for c in range(2, NUM_COLS + 1):
            ws.cell(row=current_row, column=c).border = thin_border
            ws.cell(row=current_row, column=c).fill = section_fill
        current_row += 1

        link_font_normal = Font(name="Arial", size=10, color="1155CC", underline="single")
        link_font_blue = Font(name="Arial", size=10, color="1155CC", underline="single")

        for idx, row in enumerate(data_rows):
            fill = row_fill_blue if (idx % 2 == 1) else None
            length_v = row.get("length_ft")
            width_v = row.get("width_ft")
            values = [
                row["name"], row["description"], row["address"],
                round(row["lat"], 4), round(row["lon"], 4),
                length_v if length_v else "N/A",
                width_v if width_v else "N/A",
                None,
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=v)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if fill:
                    cell.fill = fill

            ge_url = row.get("google_earth_url", "")
            ge_cell = ws.cell(row=current_row, column=8, value="View on Earth")
            ge_cell.hyperlink = ge_url
            ge_cell.font = link_font_blue if fill else link_font_normal
            ge_cell.alignment = Alignment(vertical="center", wrap_text=True)
            ge_cell.border = thin_border
            if fill:
                ge_cell.fill = fill

            current_row += 1
            total += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, total

def main():
    st.set_page_config(
        page_title="US Sports Facility Finder",
        page_icon="🏟️",
        layout="wide",
    )

    st.title("🏟️ US Sports Facility Finder")
    st.markdown("""
    Discover all sports facilities in any US city using free OpenStreetMap data.
    Pick a sport, enter a city + county, and get a formatted Excel file with all
    facilities — parks, schools, colleges, gyms, recreation centers.
    """)

    with st.sidebar:
        st.header("⚙️ Search Parameters")

        sport_choice = st.selectbox(
            "🏀 Sport",
            options=list(SPORTS_CONFIG.keys()),
            help="Select the type of facility to search for",
        )

        st.divider()
        st.subheader("📍 Location (United States)")

        city = st.text_input(
            "City",
            value="Daly City",
            help="Enter the US city name (e.g., 'Daly City', 'Alameda', 'Berkeley')",
        )

        county = st.text_input(
            "County",
            value="San Mateo County",
            help="Enter the county name (e.g., 'San Mateo County', 'Alameda County')",
        )

        state = st.text_input(
            "State",
            value="California",
            help="Enter the state name",
        )

        country = "United States"
        st.text_input("Country", value=country, disabled=True)

        st.divider()

        with st.expander("⚡ Performance / Advanced", expanded=False):
            st.markdown("**Overpass API server**")
            mirror_options = ["Auto (try all public mirrors)"] + OVERPASS_MIRRORS + ["Custom URL..."]
            mirror_choice = st.selectbox(
                "Endpoint",
                options=mirror_options,
                index=0,
                help=("If the main Overpass server blocks your requests "
                      "(common on Streamlit Cloud), try a different mirror. "
                      "The app will auto-fallback through mirrors on failure."),
            )
            if mirror_choice == "Custom URL...":
                overpass_url = st.text_input(
                    "Custom Overpass URL",
                    value="http://localhost:8080/api/interpreter",
                    help=("Use this for a self-hosted Overpass instance. "
                          "See OVERPASS_LOCAL_SETUP.md."),
                )
            elif mirror_choice == "Auto (try all public mirrors)":
                overpass_url = OVERPASS_MIRRORS[0]
            else:
                overpass_url = mirror_choice
            st.caption("💡 Local Overpass (localhost) auto-detects and uses 6 "
                        "parallel workers (vs 2 for public)")

            st.divider()
            use_cache = st.checkbox(
                "Use response cache",
                value=True,
                help=("Caches Overpass + Nominatim responses in facility_cache.db. "
                      "Second search of the same city is nearly instant."),
            )
            count, size = cache_stats()
            if count > 0:
                st.caption(f"💾 Cache: {count} entries, "
                            f"{size / 1024 / 1024:.1f} MB")
                if st.button("🗑️ Clear cache", use_container_width=True):
                    cache_clear()
                    st.success("Cache cleared. Refresh to see update.")
            else:
                st.caption("💾 Cache is empty")

        st.divider()
        run_button = st.button("🔍 Find Facilities", type="primary",
                                use_container_width=True)

    if not run_button:
        st.info("👈 Configure your search in the sidebar, then click **Find Facilities**.")
        with st.expander("ℹ️ How it works"):
            st.markdown("""
            1. **Looks up city boundaries** via OpenStreetMap Nominatim
            2. **Searches Overpass API** for sport-tagged courts/fields, parks,
               schools, colleges, gyms, recreation centers within the city bbox
            3. **Searches Nominatim** for additional facility names
            4. **Merges & deduplicates** results, groups multiple courts at the
               same facility (e.g., "Soccer Field 1", "Soccer Field 2")
            5. **Reverse geocodes** each facility to get an accurate address
            6. **Filters by city name** in the address — removes facilities that
               actually belong to neighboring cities (handles overlapping zip
               codes for irregularly-shaped cities)
            7. **Categorizes** by facility type (parks, schools, gyms, etc.)
            8. **Exports** to a formatted Excel file
            """)
        return

    if not city or not county:
        st.error("Please enter both city and county.")
        return

    sport_config = SPORTS_CONFIG[sport_choice]
    log_messages = []

    def log(msg):
        log_messages.append(msg)

    status = st.status("🍳 Cooking up your sports facility data...",
                        expanded=False)

    with status:
        st.write("Looking up city boundaries...")

        bbox = lookup_city_bbox(city, county, state, country, use_cache=use_cache)
        if not bbox:
            status.update(label="❌ City not found", state="error")
            st.error(
                f"Could not find a city named '{city}' in '{county}, {state}'.\n\n"
                "**Things to check:**\n"
                f"- Spelling of city name (try variations like '{city} City' or 'City of {city}')\n"
                "- County name should include the word 'County' (e.g., 'Alameda County')\n"
                "- For ambiguous names, try a more specific city — e.g., 'Alameda, Alameda County, CA'\n"
                "- Some cities aren't in OpenStreetMap; try a nearby major city instead"
            )
            return
        log(f"[1/6] Bbox: lat [{bbox['min_lat']:.4f}, {bbox['max_lat']:.4f}], "
            f"lon [{bbox['min_lon']:.4f}, {bbox['max_lon']:.4f}]")
        if bbox.get("match_display"):
            log(f"  Matched: {bbox['match_display'][:80]}")
        if bbox.get("polygon"):
            log(f"  Got city polygon: {len(bbox['polygon'])} points")

        is_local_overpass = ("localhost" in overpass_url or
                             "127.0.0.1" in overpass_url)

        st.write("Searching Overpass API for sports facilities...")
        t0 = time.time()
        overpass_results = fetch_overpass(bbox, sport_config, overpass_url, log,
                                           is_local=is_local_overpass,
                                           use_cache=use_cache)
        log(f"[2/6] Overpass took {time.time() - t0:.1f}s, "
            f"{len(overpass_results)} results")

        st.write("Searching Nominatim for facility names...")
        t0 = time.time()
        nominatim_results = fetch_nominatim(city, state, bbox, sport_config, log,
                                             use_cache=use_cache)
        log(f"[3/6] Nominatim took {time.time() - t0:.1f}s, "
            f"{len(nominatim_results)} results")

        all_results = overpass_results + nominatim_results

        st.write("Merging and deduplicating results...")
        merged = merge_and_deduplicate(all_results, sport_config, log)

        if not merged:
            status.update(label="No facilities found", state="error")
            st.warning("No facilities found. Try a different city or sport.")
            return

        st.write("Geocoding addresses...")
        t0 = time.time()
        reverse_geocode_all(merged, city, log, use_local_nominatim=False,
                             use_cache=use_cache)
        log(f"[5/6] Reverse geocoding took {time.time() - t0:.1f}s")

        st.write(f"Filtering to facilities inside {city}...")
        filtered, removed_list = filter_wrong_city(merged, city, state, bbox, log,
                                                      use_cache=use_cache)

        if not filtered:
            status.update(label="No facilities inside city", state="error")
            st.warning(f"No facilities found inside {city} after filtering.")
            return

        st.write("Building your Excel file...")
        categories = categorize(filtered, sport_config, log)
        excel_buffer, total = build_excel(categories, sport_config, city, state)

        status.update(label=f"✅ Done! Found {total} facilities", state="complete")

    if len(overpass_results) == 0:
        st.warning(
            "⚠️ **Overpass API was unavailable** — fell back to Nominatim only. "
            "May have missed some smaller facilities. Run locally for full data."
        )

    st.success(f"✅ Found **{total}** {sport_config['facility_label'].lower()} entries "
               f"across **{sum(1 for v in categories.values() if v)}** categories in {city}")

    col1, col2 = st.columns([1, 1])
    with col1:
        st.subheader("📊 Summary by Category")
        for cat, items in categories.items():
            if items:
                expanded = expand_to_rows(items, sport_config, cat)
                st.metric(cat, len(expanded))

    with col2:
        if removed_list:
            st.subheader(f"🚫 Removed (outside {city})")
            with st.expander(f"View {len(removed_list)} removed facilities"):
                for r in removed_list:
                    st.text(f"• {r}")

    st.subheader("📋 Preview")
    preview_data = []
    for cat, items in categories.items():
        if not items:
            continue
        for row in expand_to_rows(items, sport_config, cat):
            preview_data.append({
                "Category": cat,
                "Facility": row["name"],
                "Description": row["description"],
                "Address": row["address"],
                "Lat": round(row["lat"], 4),
                "Lon": round(row["lon"], 4),
                "Length (ft)": row.get("length_ft") or "N/A",
                "Width (ft)": row.get("width_ft") or "N/A",
                "Google Earth": row.get("google_earth_url", ""),
            })
    st.dataframe(preview_data, use_container_width=True, hide_index=True)

    with st.expander("🔍 View detailed log"):
        st.code("\n".join(log_messages), language="text")

    filename = f"{city.replace(' ', '_')}_{sport_choice.replace(' / ', '_').replace(' ', '_')}.xlsx"
    st.download_button(
        label="📥 Download Excel File",
        data=excel_buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

if __name__ == "__main__":
    main()
